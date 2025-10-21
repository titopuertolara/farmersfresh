import pandas as pd
import openpyxl

# Read the Excel file
file_path = "/home/esteban/tio cesar/Farmer's Fresh^J LLC_Profit and Loss Yearly.xlsx"

# Load workbook to see sheet names
wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)

# Read all sheets
for sheet_name in wb.sheetnames:
    print(f"\n{'='*50}")
    print(f"Sheet: {sheet_name}")
    print('='*50)
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    print(df.head(20))
    print("\nShape:", df.shape)
    print("\nColumns:", df.columns.tolist())
    print("\nData types:", df.dtypes)
